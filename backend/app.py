import os
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)
CORS(app)

# Allow large uploads
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024 * 1024  # 5 GB

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

EXCEL_FILE = "clip_times.xlsx"


# 🔹 Home Route
@app.route("/")
def home():
    return "Flask backend running!"


# 🔹 Upload Video API
@app.route("/upload", methods=["POST"])
def upload_video():
    if "video" not in request.files:
        return jsonify({"error": "No video file"}), 400

    file = request.files["video"]

    filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
    file.save(filepath)

    video_url = f"{request.host_url}video/{file.filename}"

    return jsonify({"video_url": video_url})


# 🔹 Serve Video API
@app.route("/video/<filename>")
def get_video(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


# 🔹 Save Clip Data to Excel
@app.route("/save_clip", methods=["POST"])
def save_clip():
    data = request.json
    start = data.get("start")
    end = data.get("end")

    if start is None or end is None:
        return jsonify({"error": "Missing data"}), 400

    # Create Excel file if not exists
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Start Time (sec)", "End Time (sec)"])
        wb.save(EXCEL_FILE)

    # Load and append data
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        start,
        end
    ])

    wb.save(EXCEL_FILE)

    return jsonify({"message": "Clip saved successfully"})


# 🔹 Run Server (for local)
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
